from dataclasses import replace
from time import sleep
from pyttsx3 import speak
import requests
from openpyxl import load_workbook
import bs4

wb = load_workbook('C:\\Users\\User\\Downloads\\J.A.R.V.I.S\\Stock Market\\Test Stock.xlsx')
ws = wb.active
ws['A1'].value = "No"
wb.save('C:\\Users\\User\\Downloads\\J.A.R.V.I.S\\Stock Market\\Test Stock.xlsx')
print("Axis Bank")
speak("Hello Axis Bank")
def pyautostock():
    sourceurl = requests.get('https://www.google.com/finance/quote/AXISBANK:NSE')
    gettext = bs4.BeautifulSoup(sourceurl.text, 'html.parser')
    sourcetag = gettext.find_all('div', attrs = {'class':'YMlKec fxKbKc'})
    sourcetag = str(sourcetag)
    sourcetag = sourcetag.replace('[<div class="YMlKec fxKbKc">₹', '')
    source = sourcetag.replace('</div>]', '')
    prevD = source
    sleep(60*5)
    sourceurl1 = requests.get('https://www.google.com/finance/quote/AXISBANK:NSE')
    gettext1 = bs4.BeautifulSoup(sourceurl1.text, 'html.parser')
    sourcetag1 = gettext1.find_all('div', attrs = {'class':'YMlKec fxKbKc'})
    sourcetag1 = str(sourcetag1)
    sourcetag1 = sourcetag1.replace('[<div class="YMlKec fxKbKc">₹', '')
    source1 = sourcetag1.replace('</div>]', '')
    presD = source1
    if prevD < presD:
        if ws["A1"].value == "Yes":
            speak("hello profit")
        elif ws["A1"].value == "No":
            speak("Hello Profit change")
            ws['A1'].value = "Yes"
    elif prevD > presD:
        if ws['A1'].value == "No":
            speak("hello loss")
        elif ws['A1'].value == "Yes":
            speak("hello loss change")
            ws['A1'].value = "No"
    print(prevD, presD)
    pyautostock()

pyautostock()