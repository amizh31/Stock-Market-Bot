from time import sleep
from openpyxl import load_workbook
import datetime

wb = load_workbook('C:\\Users\\User\\Downloads\\J.A.R.V.I.S\\Stock Market\\Test Stock.xlsx')
ws = wb.active
a = 0
acell = ws['A65'].value
while a<20:
    sleep(1)
    hour = str(datetime.datetime.now().strftime("%I"))
    minute = str(datetime.datetime.now().strftime("%M"))
    sec = str(datetime.datetime.now().strftime("%S"))
    month = str(datetime.datetime.now().month)
    year = str(datetime.datetime.now().year)
    date = str(datetime.datetime.now().day)
    day_1 = (int(date)+5)%7
    date = str(date)

    if day_1 == 0:
        day_1 = "Sunday"
    elif day_1 == 1:
        day_1 = "Monday"
    elif day_1 == 2:
        day_1 = "Tuesday"
    elif day_1 == 3:
        day_1 = "Wednesday"
    elif day_1 == 4:
        day_1 = "Thursday"
    elif day_1 == 5:
        day_1 = "Friday"
    elif day_1 == 6:
        day_1 = "Saturday"
    wb = load_workbook('C:\\Users\\User\\Downloads\\J.A.R.V.I.S\\Stock Market\\Test Stock.xlsx')
    ws = wb.active
    timeD = day_1+str(" ")+date+str("-")+month+str("-")+year+str(" ")+hour+str(":")+minute+str(":")+sec
    ws['A'+str(acell)+''].value = timeD
    wb.save('C:\\Users\\User\\Downloads\\J.A.R.V.I.S\\Stock Market\\Test Stock.xlsx')
    acell = acell+1
    a = a+1
ws['A'+str(acell)+''].value = acell
wb.save('C:\\Users\\User\\Downloads\\J.A.R.V.I.S\\Stock Market\\Test Stock.xlsx')
